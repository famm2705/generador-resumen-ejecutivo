from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
except ImportError as exc:
    raise SystemExit(
        "Falta la dependencia openpyxl. Instalar con: python -m pip install -r requirements.txt"
    ) from exc


SUMMARY_SHEET = "Resumen Ejecutivo PMO"
PROMPT_FILE = "Prompt.txt"
OUTPUT_SUFFIX = "_Resumen_PMO"

PALETTE = {
    "navy": "1F2937",
    "blue": "2563EB",
    "sky": "DBEAFE",
    "green": "2E7D32",
    "green_light": "E8F5E9",
    "amber": "F59E0B",
    "amber_light": "FEF3C7",
    "red": "D32F2F",
    "red_light": "FDECEC",
    "gray": "6B7280",
    "gray_light": "F3F4F6",
    "white": "FFFFFF",
    "black": "111827",
}

MONEY_COLUMNS = {"PV", "EV", "AC", "SV", "CV", "BAC", "EAC", "ETC", "VAC"}
INDEX_COLUMNS = {"SPI", "CPI", "TCPI"}


@dataclass
class SourceTable:
    sheet_name: str
    header_row: int
    data_start_row: int
    data_end_row: int
    columns: dict[str, int]
    raw_headers: dict[str, str]
    score: int


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[%/_\-.()]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("$", "").replace("%", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def safe_percent(value: Any) -> float | None:
    number = safe_float(value)
    if number is None:
        return None
    if isinstance(value, str) and "%" in value:
        return number / 100
    return number / 100 if number > 1 else number


def safe_div(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or abs(denominator) < 1e-9:
        return None
    return numerator / denominator


def as_number(value: float | None, default: float = 0.0) -> float:
    return default if value is None or not math.isfinite(value) else value


def parse_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    formats = (
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def classify_index(value: float | None) -> str:
    if value is None:
        return "N/D"
    if value >= 0.95:
        return "Verde"
    if value >= 0.85:
        return "Amarillo"
    return "Rojo"


def cost_data_anomaly(metrics: dict[str, Any]) -> bool:
    ac = as_number(metrics.get("AC"))
    ev = as_number(metrics.get("EV"))
    bac = as_number(metrics.get("BAC"))
    if ev <= 0 or bac <= 0:
        return False
    return ac <= 0 or (ac / ev < 0.05 and ac / bac < 0.02)


def classify_cpi(metrics: dict[str, Any]) -> str:
    if cost_data_anomaly(metrics):
        return "Dato atípico"
    return classify_index(metrics.get("CPI"))


def classify_tcpi(value: float | None) -> str:
    if value is None:
        return "N/D"
    if value <= 1.0:
        return "Favorable"
    if value <= 1.1:
        return "Observación"
    return "Presión"


def status_fill(status: str) -> PatternFill:
    normalized = normalize_text(status)
    if "verde" in normalized or "favorable" in normalized or "positivo" in normalized or "bajo" in normalized:
        return PatternFill("solid", fgColor=PALETTE["green_light"])
    if "amarillo" in normalized or "observacion" in normalized or "moderado" in normalized or "estable" in normalized or "atipico" in normalized or "validacion" in normalized:
        return PatternFill("solid", fgColor=PALETTE["amber_light"])
    if "rojo" in normalized or "critico" in normalized or "alto" in normalized or "presion" in normalized:
        return PatternFill("solid", fgColor=PALETTE["red_light"])
    return PatternFill("solid", fgColor=PALETTE["gray_light"])


def canonical_header(header: Any) -> str | None:
    normalized = normalize_text(header)
    if not normalized:
        return None

    if "physical" in normalized and ("complete" in normalized or "earned value" in normalized):
        return "PHYSICAL_PCT_COMPLETE"
    if ("complete" in normalized or "completado" in normalized or "avance" in normalized) and (
        "planned value" in normalized or "planificado" in normalized
    ):
        return "PLANNED_PCT_COMPLETE"

    exact = {
        "pv": "PV",
        "ev": "EV",
        "ac": "AC",
        "sv": "SV",
        "cv": "CV",
        "spi": "SPI",
        "cpi": "CPI",
        "eac": "EAC",
        "etc": "ETC",
        "bac": "BAC",
        "vac": "VAC",
        "irpc": "TCPI",
        "tcpi": "TCPI",
    }
    if normalized in exact:
        return exact[normalized]

    contains = (
        ("WBS", ("wbs", "edt")),
        ("TASK", ("nombre de tarea", "actividad", "task name", "nombre actividad", "tarea")),
        ("RESPONSIBLE", ("responsable", "owner", "recurso", "resource")),
        ("START", ("fecha inicio", "inicio", "start")),
        ("FINISH", ("fecha fin", "terminacion", "termino", "finish", "fin")),
        ("BASELINE", ("linea base", "baseline", "base line")),
        ("PCT_COMPLETE", ("% completado", "porcentaje completado", "complete", "completado", "avance")),
        ("FLOAT", ("float", "holgura")),
        ("CRITICAL_PATH", ("ruta critica", "critical path", "critica")),
        ("MILESTONE", ("hito", "milestone")),
        ("STATUS", ("estado", "status")),
    )
    for canonical, options in contains:
        if any(option in normalized for option in options):
            return canonical
    return None


def detect_source_table(workbook) -> SourceTable:
    best: SourceTable | None = None
    for ws in workbook.worksheets:
        if ws.title == SUMMARY_SHEET:
            continue
        max_scan_row = min(ws.max_row, 80)
        max_scan_col = min(ws.max_column, 60)
        for row in range(1, max_scan_row + 1):
            columns: dict[str, int] = {}
            raw_headers: dict[str, str] = {}
            metric_hits = 0
            for col in range(1, max_scan_col + 1):
                header = ws.cell(row, col).value
                canonical = canonical_header(header)
                if canonical and canonical not in columns:
                    columns[canonical] = col
                    raw_headers[canonical] = str(header)
                    if canonical in MONEY_COLUMNS or canonical in INDEX_COLUMNS:
                        metric_hits += 1
            has_task = "TASK" in columns
            score = len(columns) + metric_hits * 3 + (8 if has_task else 0)
            if metric_hits >= 4 and score >= 18:
                data_end = find_data_end(ws, row + 1, columns)
                candidate = SourceTable(ws.title, row, row + 1, data_end, columns, raw_headers, score)
                if best is None or candidate.score > best.score:
                    best = candidate
    if best is None:
        raise ValueError("No se encontró una tabla EVM válida en el workbook.")
    return best


def find_data_end(ws, data_start: int, columns: dict[str, int]) -> int:
    relevant_cols = [columns[key] for key in ("TASK", "PV", "EV", "AC", "BAC", "SPI", "CPI") if key in columns]
    last = data_start
    for row in range(data_start, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in relevant_cols):
            last = row
    return last


def read_rows(workbook, table: SourceTable) -> list[dict[str, Any]]:
    ws = workbook[table.sheet_name]
    rows: list[dict[str, Any]] = []
    for row_idx in range(table.data_start_row, table.data_end_row + 1):
        row: dict[str, Any] = {
            "__excel_row": row_idx,
            "__critical_by_red_fill": row_has_red_fill(ws, row_idx, table),
        }
        has_value = False
        for canonical, col in table.columns.items():
            value = ws.cell(row_idx, col).value
            row[canonical] = value
            if value not in (None, ""):
                has_value = True
        if not has_value:
            continue
        task_name = str(row.get("TASK") or "").strip()
        numeric_signal = any(safe_float(row.get(key)) not in (None, 0.0) for key in MONEY_COLUMNS | INDEX_COLUMNS)
        if task_name or numeric_signal:
            calculate_row_metrics(row)
            rows.append(row)
    if not rows:
        raise ValueError("La tabla detectada no contiene filas de datos.")
    return rows


def row_has_red_fill(ws, row_idx: int, table: SourceTable) -> bool:
    for col_idx in range(1, max(table.columns.values()) + 1):
        fill = ws.cell(row_idx, col_idx).fill
        if not fill or fill.fill_type != "solid":
            continue
        color = fill.fgColor
        if color.type == "rgb" and str(color.rgb or "").upper().endswith("FFC7CE"):
            return True
    return False


def is_critical_path(row: dict[str, Any]) -> bool:
    if row.get("__critical_by_red_fill"):
        return True
    value = row.get("CRITICAL_PATH")
    if isinstance(value, bool):
        return value
    return normalize_text(value) in {"si", "yes", "true", "1", "critico", "critical"}


def calculate_row_metrics(row: dict[str, Any]) -> None:
    for key in MONEY_COLUMNS | INDEX_COLUMNS:
        row[key] = safe_float(row.get(key))

    pv, ev, ac, bac = row.get("PV"), row.get("EV"), row.get("AC"), row.get("BAC")

    row["SV"] = row["SV"] if row.get("SV") is not None else delta(ev, pv)
    row["CV"] = row["CV"] if row.get("CV") is not None else delta(ev, ac)
    row["SPI"] = row["SPI"] if row.get("SPI") is not None else safe_div(ev, pv)
    row["CPI"] = row["CPI"] if row.get("CPI") is not None else safe_div(ev, ac)
    row["EAC"] = row["EAC"] if row.get("EAC") is not None else safe_div(bac, row.get("CPI"))
    row["ETC"] = row.get("ETC") if row.get("ETC") is not None else delta(row.get("EAC"), ac)
    row["VAC"] = row["VAC"] if row.get("VAC") is not None else delta(bac, row.get("EAC"))
    row["TCPI"] = row["TCPI"] if row.get("TCPI") is not None else safe_div(delta(bac, ev), delta(bac, ac))
    row["START_DATE"] = parse_date(row.get("START"))
    row["FINISH_DATE"] = parse_date(row.get("FINISH"))
    planned_progress = row.get("PLANNED_PCT_COMPLETE")
    if planned_progress is None:
        planned_progress = row.get("PCT_COMPLETE")
    row["PLANNED_PROGRESS"] = safe_percent(planned_progress)
    row["PHYSICAL_PROGRESS"] = safe_percent(row.get("PHYSICAL_PCT_COMPLETE"))
    row["IS_CRITICAL_PATH"] = is_critical_path(row)


def delta(left: float | None, right: float | None) -> float | None:
    if left is None or right is None:
        return None
    return left - right


def select_global_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    first = rows[0]
    required = sum(1 for key in ("PV", "EV", "AC", "BAC") if first.get(key) is not None)
    if required >= 3:
        return first

    candidates = [row for row in rows if as_number(row.get("BAC")) > 0 and row.get("PV") is not None and row.get("EV") is not None]
    if candidates:
        return max(candidates, key=lambda item: as_number(item.get("BAC")))

    aggregate: dict[str, Any] = {"TASK": "Agregado calculado"}
    for key in ("PV", "EV", "AC", "BAC"):
        aggregate[key] = sum(as_number(row.get(key)) for row in rows)
    calculate_row_metrics(aggregate)
    return aggregate


def row_is_analyzable(row: dict[str, Any]) -> bool:
    if not str(row.get("TASK") or "").strip():
        return False
    money_signal = any(abs(as_number(row.get(key))) > 1e-9 for key in ("PV", "EV", "AC", "BAC", "SV", "CV"))
    return money_signal or bool(row.get("IS_CRITICAL_PATH"))


def activity_risk(row: dict[str, Any]) -> dict[str, Any] | None:
    if not row_is_analyzable(row):
        return None

    spi, cpi = row.get("SPI"), row.get("CPI")
    sv, cv, vac = row.get("SV"), row.get("CV"), row.get("VAC")
    float_value = safe_float(row.get("FLOAT"))
    score = 0.0
    drivers: list[str] = []
    critical_path = bool(row.get("IS_CRITICAL_PATH"))

    if critical_path:
        score += 45
        drivers.append("ruta crítica")

    if spi is not None and as_number(row.get("PV")) > 0 and spi < 0.85:
        score += min(35, (0.85 - spi) * 60)
        drivers.append("bajo SPI")
    if cpi is not None and cpi < 0.85 and (as_number(row.get("AC")) > 0 or as_number(row.get("BAC")) > 0):
        score += min(35, (0.85 - cpi) * 60)
        drivers.append("bajo CPI")
    if float_value is not None and float_value < 0:
        score += 25
        drivers.append("holgura negativa")
    if sv is not None and sv < 0:
        score += min(20, abs(sv) / 1000)
        drivers.append("atraso")
    if cv is not None and cv < 0:
        score += min(25, abs(cv) / 1000)
        drivers.append("sobrecosto")
    if vac is not None and vac < 0:
        score += min(20, abs(vac) / 2000)
        drivers.append("presión EAC")

    if score <= 0:
        return None

    severity = round(min(100, max(1, score)), 1)
    if severity >= 70:
        label = "Crítico"
    elif severity >= 45:
        label = "Alto"
    elif severity >= 25:
        label = "Moderado"
    else:
        label = "Bajo"

    impact_value = max(
        abs(as_number(row.get("SV"))),
        abs(as_number(row.get("CV"))),
        abs(as_number(row.get("VAC")) if row.get("VAC") is not None and as_number(row.get("VAC")) < 0 else 0),
    )

    return {
        "Actividad": str(row.get("TASK") or "N/D").strip(),
        "WBS": str(row.get("WBS") or "N/D").strip(),
        "Responsable": str(row.get("RESPONSIBLE") or "N/D").strip(),
        "Inicio": row.get("START_DATE"),
        "Fin": row.get("FINISH_DATE"),
        "RutaCritica": critical_path,
        "SPI": row.get("SPI"),
        "CPI": row.get("CPI"),
        "Float": float_value if float_value is not None else "N/D",
        "Impacto": impact_value,
        "Riesgo": label,
        "Severidad": severity,
        "Drivers": ", ".join(drivers) if drivers else "desviación EVM",
        "Accion": corrective_action(drivers),
    }


def corrective_action(drivers: Iterable[str]) -> str:
    driver_set = set(drivers)
    if "ruta crítica" in driver_set:
        return "Proteger fechas, restricciones y sucesoras de la ruta crítica con seguimiento semanal."
    if "sobrecosto" in driver_set or "bajo CPI" in driver_set:
        return "Revisar costo incurrido, compromisos y productividad semanal."
    if "bajo SPI" in driver_set or "atraso" in driver_set:
        return "Actualizar plan de recuperación y proteger actividades sucesoras."
    if "holgura negativa" in driver_set:
        return "Validar ruta crítica, restricciones y fechas compromiso."
    return "Validar paquete, responsable y acción correctiva en comité semanal."


def build_activity_risks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    risks = [risk for row in rows[1:] for risk in [activity_risk(row)] if risk]
    risks.sort(key=lambda item: (item["RutaCritica"], item["Severidad"], item["Impacto"]), reverse=True)
    return risks


def build_top_risks(activity_risks: list[dict[str, Any]], global_metrics: dict[str, Any]) -> list[dict[str, Any]]:
    top: list[dict[str, Any]] = []
    for index, item in enumerate(activity_risks[:10], start=1):
        probability = min(0.95, 0.35 + item["Severidad"] / 100 * 0.55)
        consequence = "Puede deteriorar costo y cronograma del paquete si no se corrige."
        if "bajo SPI" in item["Drivers"] or "atraso" in item["Drivers"]:
            consequence = "Puede desplazar entregables sucesores y aumentar probabilidad de atraso."
        if "bajo CPI" in item["Drivers"] or "sobrecosto" in item["Drivers"]:
            consequence = "Puede consumir contingencia y deteriorar la proyección EAC."
        top.append(
            {
                "#": index,
                "Riesgo": f"Desempeño EVM adverso: {item['Drivers']}",
                "Probabilidad": probability,
                "Impacto": item["Riesgo"],
                "Severidad": item["Severidad"],
                "Actividad afectada": item["Actividad"],
                "Consecuencia": consequence,
                "Mitigacion": item["Accion"],
            }
        )

    if not top:
        spi_status = classify_index(global_metrics.get("SPI"))
        cpi_status = classify_index(global_metrics.get("CPI"))
        top.append(
            {
                "#": 1,
                "Riesgo": "Sin actividades críticas detectadas con la fuente actual",
                "Probabilidad": 0.1,
                "Impacto": "Bajo",
                "Severidad": 10,
                "Actividad afectada": "Global",
                "Consecuencia": f"Estado global SPI {spi_status} / CPI {cpi_status}.",
                "Mitigacion": "Mantener seguimiento EVM y actualizar exportación semanal.",
            }
        )
    return top


def delay_probability(global_metrics: dict[str, Any], rows: list[dict[str, Any]], activity_risks: list[dict[str, Any]]) -> tuple[float, str]:
    spi = global_metrics.get("SPI")
    if spi is None:
        base = 0.35
    elif spi >= 1.0:
        base = 0.15
    elif spi >= 0.95:
        base = 0.25
    elif spi >= 0.85:
        base = 0.45
    else:
        base = 0.7

    analyzable_count = max(1, sum(1 for row in rows[1:] if row_is_analyzable(row)))
    critical_ratio = min(0.35, len(activity_risks) / analyzable_count * 1.8)
    path_count = sum(1 for row in rows[1:] if row.get("IS_CRITICAL_PATH"))
    path_component = min(0.2, path_count / analyzable_count * 0.5)
    negative_float_count = sum(1 for row in rows[1:] if (safe_float(row.get("FLOAT")) or 0) < 0)
    float_component = min(0.15, negative_float_count / analyzable_count)
    probability = min(0.95, max(0.05, base + critical_ratio + path_component + float_component))

    if probability < 0.25:
        level = "Bajo"
    elif probability < 0.45:
        level = "Moderado"
    elif probability < 0.65:
        level = "Alto"
    else:
        level = "Crítico"
    return round(probability, 2), level


def trend_label(global_metrics: dict[str, Any], activity_risks: list[dict[str, Any]]) -> str:
    spi, cpi = global_metrics.get("SPI"), global_metrics.get("CPI")
    if spi is not None and cpi is not None and spi >= 0.95 and cpi >= 0.95 and len(activity_risks) <= 10:
        return "Tendencia positiva"
    if (spi is not None and spi < 0.85) or (cpi is not None and cpi < 0.85):
        return "Riesgo crítico"
    if (spi is not None and spi < 0.95) or (cpi is not None and cpi < 0.95):
        return "Tendencia negativa"
    return "Tendencia estable"


def forecast_finish(global_metrics: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    finish_dates = [global_metrics["FINISH_DATE"]] if global_metrics.get("FINISH_DATE") else [
        row["FINISH_DATE"] for row in rows if row.get("FINISH_DATE")
    ]
    if not finish_dates:
        return "No calculable con fecha"
    latest_finish = max(finish_dates)
    spi = global_metrics.get("SPI")
    if spi is None or spi <= 0:
        return latest_finish.strftime("%d/%m/%Y")
    if spi >= 1:
        return latest_finish.strftime("%d/%m/%Y")
    days_slip = int(round((1 / spi - 1) * 30))
    return (latest_finish + timedelta(days=max(0, days_slip))).strftime("%d/%m/%Y")


def read_prompt_text(folder: Path) -> str:
    prompt_path = folder / PROMPT_FILE
    if not prompt_path.exists():
        return ""
    try:
        return prompt_path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        return prompt_path.read_text(encoding="latin-1")


def load_env_file(folder: Path) -> Path | None:
    candidates: list[Path] = []
    explicit = os.getenv("PMO_ENV_FILE")
    if explicit:
        candidates.append(Path(explicit).expanduser())
    candidates.extend([folder / ".venv" / ".env", folder / ".env"])

    for path in candidates:
        if not path.exists() or not path.is_file():
            continue
        for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
        return path
    return None


def build_executive_narrative(
    project_name: str,
    global_metrics: dict[str, Any],
    delay: tuple[float, str],
    activity_risks: list[dict[str, Any]],
    missing_fields: list[str],
) -> list[str]:
    spi, cpi = global_metrics.get("SPI"), global_metrics.get("CPI")
    bac, eac, vac = global_metrics.get("BAC"), global_metrics.get("EAC"), global_metrics.get("VAC")
    status = overall_status(global_metrics, delay)
    risk_count = len(activity_risks)
    critical_count = sum(1 for item in activity_risks if item.get("RutaCritica"))
    primary_risk = activity_risks[0]["Actividad"] if activity_risks else "sin actividad crítica prioritaria"
    limitation = ""
    if missing_fields:
        limitation = f" La fuente no incluye {', '.join(missing_fields[:6])}; esos elementos se reportan como N/D y limitan el forecast calendario."
    cost_warning = ""
    if cost_data_anomaly(global_metrics):
        cost_warning = (
            " El AC reportado es muy bajo frente a EV y BAC; por ello CPI, EAC, ETC y VAC "
            "son matemáticamente consistentes pero no deben usarse para decisión hasta validar costos reales."
        )
    return [
        f"El proyecto {project_name} se clasifica como {status}. El desempeño global muestra SPI {format_index(spi)} y CPI {format_index(cpi)}, con avance físico {format_percent(project_physical_progress(global_metrics))} frente a avance planificado {format_percent(project_planned_progress(global_metrics))}.",
        f"Financieramente, el BAC es {format_money(bac)}, el EAC proyectado es {format_money(eac)} y la variación al cierre es {format_money(vac)}. La lectura ejecutiva indica {classify_cpi(global_metrics).lower()} en costo y {classify_index(spi).lower()} en cronograma.{cost_warning}",
        f"La probabilidad estimada de atraso es {format_percent(delay[0])} con nivel {delay[1]}. Se detectan {risk_count} actividades/paquetes con señales de riesgo y el tablero prioriza las 10 principales; {critical_count} pertenecen a la ruta crítica identificada en rojo. El primer foco es {primary_risk}.",
        f"Recomendación PMO: proteger las fechas y sucesoras de la ruta crítica, validar semanalmente las desviaciones EVM y formalizar un plan de recuperación para los paquetes con SPI/CPI bajo 0.85.{limitation}",
    ]


def enhance_narrative_with_openai(
    prompt_text: str,
    local_narrative: list[str],
    project_name: str,
    global_metrics: dict[str, Any],
    delay: tuple[float, str],
    activity_risks: list[dict[str, Any]],
    missing: list[str],
) -> tuple[list[str], str]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return local_narrative + ["Nota: OpenAI API fue solicitada, pero OPENAI_API_KEY no esta configurada."], "local"

    try:
        from openai import OpenAI
    except ImportError:
        return local_narrative + ["Nota: OpenAI API fue solicitada, pero falta instalar el paquete openai."], "local"

    payload = {
        "project_name": project_name,
        "metrics": {
            "SPI": global_metrics.get("SPI"),
            "CPI": global_metrics.get("CPI"),
            "PV": global_metrics.get("PV"),
            "EV": global_metrics.get("EV"),
            "AC": global_metrics.get("AC"),
            "BAC": global_metrics.get("BAC"),
            "EAC": global_metrics.get("EAC"),
            "ETC": global_metrics.get("ETC"),
            "VAC": global_metrics.get("VAC"),
            "SV": global_metrics.get("SV"),
            "CV": global_metrics.get("CV"),
            "TCPI": global_metrics.get("TCPI"),
        },
        "delay_probability": {"value": delay[0], "level": delay[1]},
        "top_activity_risks": activity_risks[:5],
        "missing_source_fields": missing,
        "local_narrative": local_narrative,
    }
    instructions = (
        "Eres un sistema PMO Executive Intelligence especializado en EVM. "
        "Redacta una narrativa ejecutiva en espanol para Steering Committee. "
        "No inventes datos; si falta informacion, indicalo. "
        "Devuelve exactamente 4 parrafos breves, sin bullets y sin encabezados."
    )
    if prompt_text.strip():
        instructions += "\n\nCriterios del Prompt.txt:\n" + prompt_text[:4000]

    try:
        client = OpenAI(api_key=api_key)
        response = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-5.5"),
            instructions=instructions,
            input=json.dumps(payload, ensure_ascii=False, default=str),
            max_output_tokens=900,
        )
        text = (response.output_text or "").strip()
    except Exception as exc:
        return local_narrative + [f"Nota: no fue posible usar OpenAI API ({exc})."], "local"

    paragraphs = [line.strip() for line in re.split(r"\n\s*\n", text) if line.strip()]
    if len(paragraphs) < 2:
        paragraphs = [line.strip() for line in text.splitlines() if line.strip()]
    if not paragraphs:
        return local_narrative, "local"
    return paragraphs[:4], "OpenAI API"


def overall_status(global_metrics: dict[str, Any], delay: tuple[float, str]) -> str:
    spi_status = classify_index(global_metrics.get("SPI"))
    cpi_status = classify_cpi(global_metrics)
    if cpi_status == "Dato atípico":
        return "REQUIERE VALIDACIÓN"
    if "Rojo" in (spi_status, cpi_status) or delay[1] in {"Alto", "Crítico"}:
        return "CRÍTICO"
    if "Amarillo" in (spi_status, cpi_status) or delay[1] == "Moderado":
        return "EN OBSERVACIÓN"
    return "CONTROLADO"


def format_money(value: Any) -> str:
    number = safe_float(value)
    if number is None:
        return "N/D"
    return f"${number:,.0f}"


def format_index(value: Any) -> str:
    number = safe_float(value)
    if number is None:
        return "N/D"
    return f"{number:.2f}"


def format_percent(value: Any) -> str:
    number = safe_float(value)
    if number is None:
        return "N/D"
    return f"{number:.0%}"


def project_physical_progress(metrics: dict[str, Any]) -> float | None:
    if metrics.get("PHYSICAL_PROGRESS") is not None:
        return metrics.get("PHYSICAL_PROGRESS")
    return safe_div(metrics.get("EV"), metrics.get("BAC"))


def project_planned_progress(metrics: dict[str, Any]) -> float | None:
    if metrics.get("PLANNED_PROGRESS") is not None:
        return metrics.get("PLANNED_PROGRESS")
    return safe_div(metrics.get("PV"), metrics.get("BAC"))


def missing_fields(table: SourceTable, rows: list[dict[str, Any]]) -> list[str]:
    expected = {
        "WBS": "WBS",
        "RESPONSIBLE": "Responsable",
        "START": "Fecha inicio",
        "FINISH": "Fecha fin",
        "PCT_COMPLETE": "% completado",
        "FLOAT": "Float/Holgura",
        "CRITICAL_PATH": "Ruta crítica",
        "MILESTONE": "Hitos",
        "STATUS": "Estado",
    }
    missing = [label for key, label in expected.items() if key not in table.columns]
    if any(row.get("IS_CRITICAL_PATH") for row in rows):
        missing = [label for label in missing if label != "Ruta crítica"]
    if "PHYSICAL_PCT_COMPLETE" in table.columns or "PLANNED_PCT_COMPLETE" in table.columns:
        missing = [label for label in missing if label != "% completado"]
    return missing


def write_dashboard(
    workbook,
    table: SourceTable,
    rows: list[dict[str, Any]],
    prompt_text: str,
    use_openai: bool = False,
) -> None:
    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]
    ws = workbook.create_sheet(SUMMARY_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    global_row = select_global_row(rows)
    project_name = str(global_row.get("TASK") or "Proyecto EVM").strip()
    activity_risks = build_activity_risks(rows)
    top_risks = build_top_risks(activity_risks, global_row)
    delay = delay_probability(global_row, rows, activity_risks)
    trend = trend_label(global_row, activity_risks)
    finish_forecast = forecast_finish(global_row, rows)
    missing = missing_fields(table, rows)
    narrative = build_executive_narrative(project_name, global_row, delay, activity_risks, missing)
    narrative_source = "local"
    if use_openai:
        narrative, narrative_source = enhance_narrative_with_openai(
            prompt_text=prompt_text,
            local_narrative=narrative,
            project_name=project_name,
            global_metrics=global_row,
            delay=delay,
            activity_risks=activity_risks,
            missing=missing,
        )

    setup_page(ws)
    write_title(ws, project_name, table, rows, prompt_text, narrative_source)
    write_kpi_cards(ws, global_row, delay, trend)
    write_narrative(ws, narrative)
    write_kpi_table(ws, global_row, delay, trend, finish_forecast)
    write_forecast_table(ws, global_row, trend, delay, finish_forecast, missing)
    write_activity_table(ws, activity_risks)
    write_risk_table(ws, top_risks)
    write_milestones_or_limits(ws, rows, missing, table)
    write_visualizations(ws, rows, global_row, activity_risks, top_risks)
    apply_conditional_formatting(ws)
    apply_borders(ws)


def setup_page(ws) -> None:
    widths = {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 28,
        "I": 22,
        "J": 22,
        "K": 24,
        "L": 24,
        "M": 3,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
        "S": 16,
        "T": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col] = ColumnDimension(ws, min=ws[col + "1"].column, max=ws[col + "1"].column, width=width)
    for row in range(1, 158):
        ws.row_dimensions[row].height = 22
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def write_title(
    ws,
    project_name: str,
    table: SourceTable,
    rows: list[dict[str, Any]],
    prompt_text: str,
    narrative_source: str,
) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = "RESUMEN EJECUTIVO PMO - Earned Value Management"
    ws["A1"].font = Font(color=PALETTE["white"], bold=True, size=18)
    ws["A1"].fill = PatternFill("solid", fgColor=PALETTE["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:L2")
    prompt_status = "Prompt.txt detectado" if prompt_text.strip() else "Sin Prompt.txt"
    ws["A2"] = (
        f"Proyecto: {project_name} | Fuente: {table.sheet_name}!A{table.header_row}:"
        f"{get_column_letter(max(table.columns.values()))}{table.data_end_row} | "
        f"Filas: {len(rows)} | Generado: {datetime.now():%d/%m/%Y %H:%M} | "
        f"{prompt_status} | Narrativa: {narrative_source}"
    )
    ws["A2"].font = Font(color=PALETTE["white"], size=10)
    ws["A2"].fill = PatternFill("solid", fgColor=PALETTE["navy"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")


def write_kpi_cards(ws, metrics: dict[str, Any], delay: tuple[float, str], trend: str) -> None:
    cpi_status = classify_cpi(metrics)
    cards = [
        ("A4:C8", "SPI Global", format_index(metrics.get("SPI")), classify_index(metrics.get("SPI")), "EV / PV"),
        ("D4:F8", "CPI Global", format_index(metrics.get("CPI")), cpi_status, "Validar AC" if cost_data_anomaly(metrics) else "EV / AC"),
        ("G4:I8", "EAC Forecast", format_money(metrics.get("EAC")), cpi_status, "No confiable hasta validar AC" if cost_data_anomaly(metrics) else "BAC / CPI"),
        ("J4:L8", "Riesgo de Atraso", format_percent(delay[0]), delay[1], trend),
    ]
    for cell_range, title, value, status, subtitle in cards:
        ws.merge_cells(cell_range)
        top_left = cell_range.split(":")[0]
        cell = ws[top_left]
        cell.value = f"{title}\n{value}\n{status} | {subtitle}"
        cell.font = Font(color=PALETTE["black"], bold=True, size=13)
        cell.fill = status_fill(status)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
    for row in range(4, 9):
        ws.row_dimensions[row].height = 26


def section_header(ws, row: int, title: str, end_col: str = "L") -> None:
    ws.merge_cells(f"A{row}:{end_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=PALETTE["blue"])
    cell.font = Font(color=PALETTE["white"], bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def table_header(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(color=PALETTE["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def merged_table_header(ws, row: int, specs: list[tuple[int, int, str]]) -> None:
    for start_col, end_col, header in specs:
        if start_col != end_col:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row, start_col)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(color=PALETTE["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def merged_value(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    value: Any,
    fill: PatternFill | None = None,
    bold: bool = False,
    number_format: str | None = None,
    horizontal: str = "left",
) -> None:
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col)
    cell.value = value
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    if fill is not None:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    if number_format:
        cell.number_format = number_format


def write_narrative(ws, narrative: list[str]) -> None:
    section_header(ws, 10, "Resumen ejecutivo automático")
    ws.merge_cells("A11:L16")
    ws["A11"] = "\n\n".join(narrative)
    ws["A11"].fill = PatternFill("solid", fgColor=PALETTE["gray_light"])
    ws["A11"].font = Font(color=PALETTE["black"], size=10)
    ws["A11"].alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(11, 17):
        ws.row_dimensions[row].height = 42


def write_kpi_table(ws, metrics: dict[str, Any], delay: tuple[float, str], trend: str, finish_forecast: str) -> None:
    section_header(ws, 18, "KPI globales y semáforos")
    headers = ["KPI", "Valor", "Estado", "Interpretación", "KPI", "Valor", "Estado", "Interpretación"]
    table_header(ws, 19, headers)
    rows = [
        ("SPI Global", metrics.get("SPI"), classify_index(metrics.get("SPI")), "EV / PV", "SV", metrics.get("SV"), "Adelanto" if as_number(metrics.get("SV")) >= 0 else "Atraso", "EV - PV"),
        ("CPI Global", metrics.get("CPI"), classify_cpi(metrics), "Validar AC antes de interpretar", "CV", metrics.get("CV"), "Dato atípico" if cost_data_anomaly(metrics) else ("Ahorro" if as_number(metrics.get("CV")) >= 0 else "Sobrecosto"), "EV - AC"),
        ("BAC", metrics.get("BAC"), "", "Presupuesto total", "EAC", metrics.get("EAC"), classify_cpi(metrics), "No confiable hasta validar AC" if cost_data_anomaly(metrics) else "BAC / CPI"),
        ("ETC", metrics.get("ETC"), "", "EAC - AC", "VAC", metrics.get("VAC"), "Favorable" if as_number(metrics.get("VAC")) >= 0 else "Presión", "BAC - EAC"),
        ("% Avance físico", project_physical_progress(metrics), "", "Physical % Complete", "% Avance planificado", project_planned_progress(metrics), "", "% Complete / Planned Value"),
        ("TCPI / IRPC", metrics.get("TCPI"), classify_tcpi(metrics.get("TCPI")), "(BAC-EV)/(BAC-AC)", "Forecast terminación", finish_forecast, trend, "Basado en fechas/SPI"),
        ("Riesgo de atraso", delay[0], delay[1], "Modelo EVM + riesgos", "Estado general", overall_status(metrics, delay), overall_status(metrics, delay), "Lectura PMO"),
    ]
    for idx, row_values in enumerate(rows, start=20):
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(idx, col)
            cell.value = value
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in {2, 6} and isinstance(value, (int, float)):
                apply_number_format(cell, row_values[col - 2])
            if col in {3, 7}:
                cell.fill = status_fill(str(value))
                cell.font = Font(bold=True)


def write_forecast_table(
    ws,
    metrics: dict[str, Any],
    trend: str,
    delay: tuple[float, str],
    finish_forecast: str,
    missing: list[str],
) -> None:
    section_header(ws, 29, "Forecast predictivo")
    merged_table_header(
        ws,
        30,
        [
            (1, 2, "Forecast"),
            (3, 4, "Resultado"),
            (5, 8, "Riesgo / lectura ejecutiva"),
            (9, 12, "Base de cálculo"),
        ],
    )
    rows = [
        ("SPI esperado al cierre", metrics.get("SPI"), trend, "SPI actual por falta de serie histórica"),
        ("CPI esperado al cierre", metrics.get("CPI"), trend, "CPI actual por falta de serie histórica"),
        (
            "Forecast financiero",
            metrics.get("EAC"),
            "Validar AC: forecast no confiable"
            if cost_data_anomaly(metrics)
            else ("Favorable vs BAC" if as_number(metrics.get("VAC")) >= 0 else "Presión vs BAC"),
            "EAC calculado con CPI; sujeto a calidad de AC",
        ),
        ("Forecast de terminación", finish_forecast, "Limitado por fechas fuente" if "Fecha fin" in missing else trend, "Fecha fin + SPI"),
        ("Probabilidad estimada de atraso", delay[0], delay[1], "SPI + concentración de actividades críticas"),
        ("Limitaciones de fuente", ", ".join(missing) if missing else "Sin campos faltantes relevantes", "No inventar información", "Detección automática de columnas"),
    ]
    for idx, row_values in enumerate(rows, start=31):
        ws.row_dimensions[idx].height = 34
        merged_value(ws, idx, 1, 2, row_values[0])
        number_format = None
        if isinstance(row_values[1], (int, float)):
            name = normalize_text(row_values[0])
            if "probabilidad" in name:
                number_format = "0%"
            elif any(token in name for token in ("spi", "cpi")):
                number_format = "0.00"
            elif "financiero" in name:
                number_format = '$#,##0;[Red]($#,##0)'
        merged_value(ws, idx, 3, 4, row_values[1], number_format=number_format)
        merged_value(ws, idx, 5, 8, row_values[2], fill=status_fill(str(row_values[2])), bold=True)
        merged_value(ws, idx, 9, 12, row_values[3])


def write_activity_table(ws, activity_risks: list[dict[str, Any]]) -> None:
    section_header(ws, 39, "Actividades prioritarias y ruta crítica")
    merged_table_header(
        ws,
        40,
        [
            (1, 2, "Actividad / Paquete"),
            (3, 3, "WBS"),
            (4, 4, "Inicio"),
            (5, 5, "Fin"),
            (6, 6, "SPI"),
            (7, 7, "CPI"),
            (8, 8, "Ruta crítica"),
            (9, 9, "Riesgo"),
            (10, 10, "Driver"),
            (11, 12, "Acción recomendada"),
        ],
    )
    if not activity_risks:
        ws.merge_cells("A41:L41")
        ws["A41"] = "No se detectaron actividades prioritarias por ruta crítica o desviaciones EVM."
        return
    for idx, item in enumerate(activity_risks[:10], start=41):
        ws.row_dimensions[idx].height = 38
        merged_value(ws, idx, 1, 2, item["Actividad"])
        merged_value(ws, idx, 3, 3, item["WBS"], horizontal="center")
        merged_value(ws, idx, 4, 4, item["Inicio"], number_format="dd/mm/yyyy", horizontal="center")
        merged_value(ws, idx, 5, 5, item["Fin"], number_format="dd/mm/yyyy", horizontal="center")
        merged_value(ws, idx, 6, 6, item["SPI"], number_format="0.00", horizontal="center")
        merged_value(ws, idx, 7, 7, item["CPI"], number_format="0.00", horizontal="center")
        merged_value(
            ws,
            idx,
            8,
            8,
            "Sí" if item["RutaCritica"] else "No",
            fill=PatternFill("solid", fgColor=PALETTE["red_light"]) if item["RutaCritica"] else None,
            bold=item["RutaCritica"],
            horizontal="center",
        )
        merged_value(ws, idx, 9, 9, item["Riesgo"], fill=status_fill(str(item["Riesgo"])), bold=True, horizontal="center")
        merged_value(ws, idx, 10, 10, item["Drivers"])
        merged_value(ws, idx, 11, 12, item["Accion"])


def write_risk_table(ws, top_risks: list[dict[str, Any]]) -> None:
    section_header(ws, 53, "Top riesgos del proyecto")
    merged_table_header(
        ws,
        54,
        [
            (1, 1, "#"),
            (2, 3, "Riesgo"),
            (4, 4, "Probabilidad"),
            (5, 5, "Impacto"),
            (6, 6, "Severidad"),
            (7, 8, "Actividad afectada"),
            (9, 10, "Consecuencia"),
            (11, 12, "Mitigación recomendada"),
        ],
    )
    for idx, risk in enumerate(top_risks[:10], start=55):
        ws.row_dimensions[idx].height = 44
        merged_value(ws, idx, 1, 1, risk["#"], horizontal="center")
        merged_value(ws, idx, 2, 3, risk["Riesgo"])
        merged_value(ws, idx, 4, 4, risk["Probabilidad"], number_format="0%", horizontal="center")
        merged_value(ws, idx, 5, 5, risk["Impacto"], fill=status_fill(str(risk["Impacto"])), bold=True, horizontal="center")
        merged_value(ws, idx, 6, 6, risk["Severidad"], fill=status_fill(str(risk["Severidad"])), bold=True, horizontal="center")
        merged_value(ws, idx, 7, 8, risk["Actividad afectada"])
        merged_value(ws, idx, 9, 10, risk["Consecuencia"])
        merged_value(ws, idx, 11, 12, risk["Mitigacion"])


def write_milestones_or_limits(ws, rows: list[dict[str, Any]], missing: list[str], table: SourceTable) -> None:
    section_header(ws, 68, "Hitos próximos y validación de fuente")
    milestone_rows = upcoming_milestones(rows, table)
    if milestone_rows:
        merged_table_header(
            ws,
            69,
            [
                (1, 3, "Hito"),
                (4, 4, "Fecha compromiso"),
                (5, 6, "Estado actual"),
                (7, 8, "Riesgo de incumplimiento"),
                (9, 10, "Dependencias"),
                (11, 12, "Recomendación"),
            ],
        )
        for idx, row in enumerate(milestone_rows, start=70):
            ws.row_dimensions[idx].height = 38
            merged_value(ws, idx, 1, 3, row["Hito"])
            merged_value(ws, idx, 4, 4, row["Fecha"], number_format="dd/mm/yyyy", horizontal="center")
            merged_value(ws, idx, 5, 6, row["Estado"])
            merged_value(ws, idx, 7, 8, row["Riesgo"], fill=status_fill(str(row["Riesgo"])), bold=True, horizontal="center")
            merged_value(ws, idx, 9, 10, row["Dependencias"])
            merged_value(ws, idx, 11, 12, row["Recomendacion"])
        return

    merged_table_header(
        ws,
        69,
        [
            (1, 2, "Campo solicitado"),
            (3, 6, "Resultado de detección"),
            (7, 8, "Lectura PMO"),
            (9, 12, "Recomendación"),
        ],
    )
    available = []
    unavailable = []
    for label, is_available in (
        ("WBS", "WBS" in table.columns),
        ("Fechas", "START" in table.columns and "FINISH" in table.columns),
        ("Ruta crítica por color rojo", any(row.get("IS_CRITICAL_PATH") for row in rows)),
        ("Responsable", "RESPONSIBLE" in table.columns),
        ("Float/Holgura", "FLOAT" in table.columns),
        ("Hitos", "MILESTONE" in table.columns),
        ("Estado", "STATUS" in table.columns),
    ):
        (available if is_available else unavailable).append(label)

    validation = [
        ("Tabla de datos", f"Detectada en {table.sheet_name}!A{table.header_row}:{get_column_letter(max(table.columns.values()))}{table.data_end_row}", "OK", "Mantener encabezados consistentes."),
        ("Encabezados EVM", ", ".join(table.raw_headers.values()), "OK", "Exportar desde MS Project con los campos EVM requeridos."),
        (
            "Campos ampliados",
            f"Disponibles: {', '.join(available)}. Faltan: {', '.join(unavailable) if unavailable else 'ninguno'}.",
            "Limitación" if unavailable else "OK",
            "Mantener WBS, Start, Finish, porcentajes y formato rojo; agregar los campos aún faltantes.",
        ),
        ("Cálculos recalculados", "SPI, CPI, SV, CV, EAC, ETC, VAC, TCPI", "OK", "Validar contra línea base y cierre contable."),
    ]
    for idx, values in enumerate(validation, start=70):
        ws.row_dimensions[idx].height = 42
        merged_value(ws, idx, 1, 2, values[0])
        merged_value(ws, idx, 3, 6, values[1])
        merged_value(ws, idx, 7, 8, values[2], fill=status_fill(str(values[2])), bold=True, horizontal="center")
        merged_value(ws, idx, 9, 12, values[3])


def write_visualizations(
    ws,
    rows: list[dict[str, Any]],
    metrics: dict[str, Any],
    activity_risks: list[dict[str, Any]],
    top_risks: list[dict[str, Any]],
) -> None:
    section_header(ws, 78, "Visualizaciones ejecutivas", "L")
    ws.merge_cells("A79:L79")
    ws["A79"] = (
        "10 graficos embebidos como imagenes PNG para evitar fallas de carga de charts nativos. "
        "Cuando la fuente no trae fechas o serie historica, las tendencias usan la secuencia de actividades exportada."
    )
    ws["A79"].fill = PatternFill("solid", fgColor=PALETTE["gray_light"])
    ws["A79"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[79].height = 34

    try:
        chart_buffers = build_chart_images(rows, metrics, activity_risks, top_risks)
    except Exception as exc:
        ws.merge_cells("A80:L82")
        ws["A80"] = f"No fue posible generar las imagenes de graficos: {exc}"
        ws["A80"].fill = PatternFill("solid", fgColor=PALETTE["red_light"])
        ws["A80"].alignment = Alignment(wrap_text=True, vertical="center")
        return

    anchors = ["A81", "G81", "A97", "G97", "A113", "G113", "A129", "G129", "A145", "G145"]
    ws._pmo_image_buffers = chart_buffers
    for buffer, anchor in zip(chart_buffers, anchors):
        image = ExcelImage(buffer)
        image.width = 500
        image.height = 285
        ws.add_image(image, anchor)


def build_chart_images(
    rows: list[dict[str, Any]],
    metrics: dict[str, Any],
    activity_risks: list[dict[str, Any]],
    top_risks: list[dict[str, Any]],
) -> list[BytesIO]:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "axes.titlesize": 10,
            "axes.labelsize": 8,
            "xtick.labelsize": 7,
            "ytick.labelsize": 7,
            "legend.fontsize": 7,
        }
    )

    money_formatter = FuncFormatter(lambda value, _: f"${value/1000:,.0f}k")
    money_detail_formatter = FuncFormatter(lambda value, _: f"${value:,.0f}")
    index_formatter = FuncFormatter(lambda value, _: f"{value:.2f}")
    analyzable = [row for row in rows[1:] if row_is_analyzable(row)]
    buffers: list[BytesIO] = []

    def finish(fig) -> BytesIO:
        buffer = BytesIO()
        fig.tight_layout()
        fig.savefig(buffer, format="png", dpi=130, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buffer.seek(0)
        return buffer

    def base_fig(title: str):
        fig, ax = plt.subplots(figsize=(5.2, 3.0))
        fig.patch.set_facecolor("white")
        ax.set_title(title, loc="left", fontweight="bold", color=f"#{PALETTE['navy']}")
        ax.grid(axis="y", color="#E5E7EB", linewidth=0.8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        return fig, ax

    def no_data(title: str, message: str = "Sin datos suficientes") -> BytesIO:
        fig, ax = base_fig(title)
        ax.text(0.5, 0.5, message, ha="center", va="center", transform=ax.transAxes, color=f"#{PALETTE['gray']}")
        ax.set_xticks([])
        ax.set_yticks([])
        return finish(fig)

    # 1. EV vs PV vs AC
    fig, ax = base_fig("1. EV vs PV vs AC")
    labels = ["PV", "EV", "AC"]
    values = [as_number(metrics.get(key)) for key in labels]
    ax.bar(labels, values, color=[f"#{PALETTE['blue']}", f"#{PALETTE['green']}", f"#{PALETTE['amber']}"])
    ax.yaxis.set_major_formatter(money_formatter)
    buffers.append(finish(fig))

    # 2. Curva S by exported sequence
    curve_rows = [row for row in analyzable if any(as_number(row.get(key)) > 0 for key in ("PV", "EV", "AC"))][:80]
    if curve_rows:
        fig, ax = base_fig("2. Curva S EVM - secuencia")
        x = list(range(1, len(curve_rows) + 1))
        for key, color in (("PV", "blue"), ("EV", "green"), ("AC", "amber")):
            total = 0.0
            y = []
            for row in curve_rows:
                total += max(0.0, as_number(row.get(key)))
                y.append(total)
            ax.plot(x, y, linewidth=1.8, label=key, color=f"#{PALETTE[color]}")
        ax.yaxis.set_major_formatter(money_formatter)
        ax.set_xlabel("Secuencia de actividades")
        ax.legend(loc="upper left")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("2. Curva S EVM - secuencia"))

    # 3. SPI vs CPI
    scatter_rows = [row for row in analyzable if row.get("SPI") is not None and row.get("CPI") is not None][:120]
    if scatter_rows:
        fig, ax = base_fig("3. SPI vs CPI")
        x = [as_number(row.get("SPI")) for row in scatter_rows]
        y = [as_number(row.get("CPI")) for row in scatter_rows]
        ax.scatter(x, y, s=22, alpha=0.75, color=f"#{PALETTE['blue']}")
        ax.axvline(0.95, color=f"#{PALETTE['red']}", linestyle="--", linewidth=1)
        ax.axhline(0.95, color=f"#{PALETTE['red']}", linestyle="--", linewidth=1)
        ax.set_xlabel("SPI")
        ax.set_ylabel("CPI")
        ax.xaxis.set_major_formatter(index_formatter)
        ax.yaxis.set_major_formatter(index_formatter)
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("3. SPI vs CPI"))

    # 4. Forecast EAC with separate scales so small values remain visible
    fig, (ax, ax_detail) = plt.subplots(1, 2, figsize=(5.2, 3.0), gridspec_kw={"width_ratios": [1.2, 1]})
    fig.patch.set_facecolor("white")
    fig.suptitle("4. Forecast financiero", x=0.02, ha="left", fontweight="bold", color=f"#{PALETTE['navy']}", fontsize=10)
    ax.bar(["BAC", "VAC"], [as_number(metrics.get("BAC")), as_number(metrics.get("VAC"))], color=[f"#{PALETTE['blue']}", f"#{PALETTE['gray']}"])
    ax_detail.bar(["EAC", "ETC"], [as_number(metrics.get("EAC")), as_number(metrics.get("ETC"))], color=[f"#{PALETTE['green']}", f"#{PALETTE['amber']}"])
    for axis in (ax, ax_detail):
        axis.grid(axis="y", color="#E5E7EB", linewidth=0.8)
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)
    ax.yaxis.set_major_formatter(money_formatter)
    ax_detail.yaxis.set_major_formatter(money_detail_formatter)
    if cost_data_anomaly(metrics):
        ax_detail.text(0.5, 0.95, "Validar AC", transform=ax_detail.transAxes, ha="center", va="top", color=f"#{PALETTE['red']}", fontweight="bold")
    buffers.append(finish(fig))

    # 5. Top risks
    risk_labels = [truncate_label(str(risk["Actividad afectada"]), 24) for risk in top_risks[:8]]
    risk_values = [as_number(risk["Severidad"]) for risk in top_risks[:8]]
    if risk_labels:
        fig, ax = base_fig("5. Riesgos principales")
        ax.barh(list(reversed(risk_labels)), list(reversed(risk_values)), color=f"#{PALETTE['red']}")
        ax.set_xlim(0, 100)
        ax.set_xlabel("Severidad")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("5. Riesgos principales"))

    # 6. Critical path durations from the red source rows
    critical_items = [
        item for item in activity_risks
        if item.get("RutaCritica") and item.get("Inicio") and item.get("Fin")
    ][:8]
    if critical_items:
        fig, ax = base_fig("6. Ruta crítica - duración")
        labels = [truncate_label(item["Actividad"], 24) for item in critical_items]
        durations = [max(0, (item["Fin"] - item["Inicio"]).days) for item in critical_items]
        ax.barh(list(reversed(labels)), list(reversed(durations)), color=f"#{PALETTE['red']}")
        ax.set_xlabel("Duración calendario (días)")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("6. Ruta crítica - duración"))

    # 7. Deviation distribution
    deviations = [as_number(row.get("SV")) for row in analyzable if row.get("SV") is not None]
    cost_deviations = [as_number(row.get("CV")) for row in analyzable if row.get("CV") is not None]
    if deviations or cost_deviations:
        fig, ax = base_fig("7. Distribucion de desviaciones")
        if deviations:
            ax.hist(deviations, bins=16, alpha=0.65, label="SV", color=f"#{PALETTE['blue']}")
        if cost_deviations:
            ax.hist(cost_deviations, bins=16, alpha=0.55, label="CV", color=f"#{PALETTE['green']}")
        ax.xaxis.set_major_formatter(money_formatter)
        ax.legend(loc="upper right")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("7. Distribucion de desviaciones"))

    # 8. Performance trend by sequence
    trend_rows = [row for row in analyzable if row.get("SPI") is not None or row.get("CPI") is not None][:80]
    if trend_rows:
        fig, ax = base_fig("8. Tendencia de desempeno")
        x = list(range(1, len(trend_rows) + 1))
        spi = [as_number(row.get("SPI"), float("nan")) for row in trend_rows]
        cpi = [as_number(row.get("CPI"), float("nan")) for row in trend_rows]
        ax.plot(x, spi, label="SPI", color=f"#{PALETTE['blue']}", linewidth=1.5)
        ax.plot(x, cpi, label="CPI", color=f"#{PALETTE['green']}", linewidth=1.5)
        ax.axhline(0.95, color=f"#{PALETTE['red']}", linestyle="--", linewidth=1)
        ax.set_xlabel("Secuencia de actividades")
        ax.yaxis.set_major_formatter(index_formatter)
        ax.legend(loc="upper right")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("8. Tendencia de desempeno"))

    # 9. Schedule variance
    negative_sv = sorted(
        [row for row in analyzable if row.get("SV") is not None and as_number(row.get("SV")) < 0],
        key=lambda row: as_number(row.get("SV")),
    )[:8]
    if negative_sv:
        fig, ax = base_fig("9. Variacion cronograma")
        labels = [truncate_label(str(row.get("TASK")), 24) for row in negative_sv]
        values = [abs(as_number(row.get("SV"))) for row in negative_sv]
        ax.barh(list(reversed(labels)), list(reversed(values)), color=f"#{PALETTE['red']}")
        ax.xaxis.set_major_formatter(money_formatter)
        ax.set_xlabel("SV negativo")
        buffers.append(finish(fig))
    else:
        buffers.append(no_data("9. Variacion cronograma", "Sin SV negativo"))

    # 10. Cost variance: show savings when no negative CV exists
    negative_cv = sorted(
        [row for row in analyzable if row.get("CV") is not None and as_number(row.get("CV")) < 0],
        key=lambda row: as_number(row.get("CV")),
    )[:8]
    if negative_cv:
        fig, ax = base_fig("10. Variacion costo")
        labels = [truncate_label(str(row.get("TASK")), 24) for row in negative_cv]
        values = [abs(as_number(row.get("CV"))) for row in negative_cv]
        ax.barh(list(reversed(labels)), list(reversed(values)), color=f"#{PALETTE['red']}")
        ax.xaxis.set_major_formatter(money_formatter)
        ax.set_xlabel("CV negativo")
        buffers.append(finish(fig))
    else:
        positive_cv = sorted(
            [row for row in analyzable if row.get("CV") is not None and as_number(row.get("CV")) > 0],
            key=lambda row: as_number(row.get("CV")),
            reverse=True,
        )[:8]
        if positive_cv:
            fig, ax = base_fig("10. Variación costo positiva")
            labels = [truncate_label(str(row.get("TASK")), 24) for row in positive_cv]
            values = [as_number(row.get("CV")) for row in positive_cv]
            ax.barh(list(reversed(labels)), list(reversed(values)), color=f"#{PALETTE['green']}")
            ax.xaxis.set_major_formatter(money_formatter)
            ax.set_xlabel("CV positivo reportado")
            if cost_data_anomaly(metrics):
                ax.text(0.98, 0.03, "Interpretar tras validar AC", transform=ax.transAxes, ha="right", color=f"#{PALETTE['red']}", fontsize=7)
            buffers.append(finish(fig))
        else:
            buffers.append(no_data("10. Variación costo", "Sin variación de costo disponible"))

    return buffers


def truncate_label(value: str, max_len: int) -> str:
    value = re.sub(r"\s+", " ", value or "N/D").strip()
    if len(value) <= max_len:
        return value
    return value[: max_len - 1] + "..."


def upcoming_milestones(rows: list[dict[str, Any]], table: SourceTable) -> list[dict[str, Any]]:
    if "MILESTONE" not in table.columns or "FINISH" not in table.columns:
        return []
    now = datetime.now()
    horizon = now + timedelta(days=14)
    milestones: list[dict[str, Any]] = []
    for row in rows:
        is_milestone = bool(row.get("MILESTONE"))
        finish = row.get("FINISH_DATE")
        if is_milestone and finish and now <= finish <= horizon:
            spi = row.get("SPI")
            risk = classify_index(spi)
            milestones.append(
                {
                    "Hito": str(row.get("TASK") or "N/D"),
                    "Fecha": finish,
                    "Estado": str(row.get("STATUS") or "N/D"),
                    "Riesgo": risk,
                    "Dependencias": "N/D",
                    "Recomendacion": "Confirmar restricciones y responsables de cierre.",
                }
            )
    milestones.sort(key=lambda item: item["Fecha"])
    return milestones[:5]


def write_chart_data(ws, metrics: dict[str, Any], top_risks: list[dict[str, Any]]) -> None:
    start = 84
    ws[f"N{start}"] = "Serie"
    ws[f"O{start}"] = "PV"
    ws[f"P{start}"] = "EV"
    ws[f"Q{start}"] = "AC"
    rows = [
        ("Global", metrics.get("PV"), metrics.get("EV"), metrics.get("AC")),
        ("Forecast", metrics.get("BAC"), metrics.get("EAC"), metrics.get("ETC")),
        ("Variación", metrics.get("SV"), metrics.get("CV"), metrics.get("VAC")),
    ]
    for idx, values in enumerate(rows, start=start + 1):
        for offset, value in enumerate(values):
            ws.cell(idx, 14 + offset).value = value

    ws[f"N{start + 6}"] = "Índice"
    ws[f"O{start + 6}"] = "Valor"
    ws[f"P{start + 6}"] = "Umbral"
    index_rows = [("SPI", metrics.get("SPI"), 0.95), ("CPI", metrics.get("CPI"), 0.95), ("TCPI", metrics.get("TCPI"), 1.0)]
    for idx, values in enumerate(index_rows, start=start + 7):
        for offset, value in enumerate(values):
            ws.cell(idx, 14 + offset).value = value

    ws[f"N{start + 12}"] = "Riesgo"
    ws[f"O{start + 12}"] = "Severidad"
    for idx, risk in enumerate(top_risks[:10], start=start + 13):
        ws.cell(idx, 14).value = str(risk["Actividad afectada"])[:28]
        ws.cell(idx, 15).value = risk["Severidad"]

    for col in range(14, 18):
        ws.column_dimensions[get_column_letter(col)].hidden = True


def add_charts(ws) -> None:
    data_row = 84
    chart1 = BarChart()
    chart1.title = "EV vs PV vs AC"
    chart1.y_axis.title = "Valor"
    chart1.x_axis.title = "Serie"
    chart1.height = 7
    chart1.width = 12
    data = Reference(ws, min_col=15, max_col=17, min_row=data_row, max_row=data_row + 3)
    cats = Reference(ws, min_col=14, min_row=data_row + 1, max_row=data_row + 3)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.legend.position = "b"
    ws.add_chart(chart1, "J20")

    chart2 = LineChart()
    chart2.title = "SPI / CPI / TCPI"
    chart2.y_axis.title = "Índice"
    chart2.height = 7
    chart2.width = 12
    idx_row = data_row + 6
    data = Reference(ws, min_col=15, max_col=16, min_row=idx_row, max_row=idx_row + 3)
    cats = Reference(ws, min_col=14, min_row=idx_row + 1, max_row=idx_row + 3)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.legend.position = "b"
    ws.add_chart(chart2, "J36")

    chart3 = BarChart()
    chart3.title = "Top Riesgos por Severidad"
    chart3.y_axis.title = "Severidad"
    chart3.height = 8
    chart3.width = 12
    risk_row = data_row + 12
    data = Reference(ws, min_col=15, min_row=risk_row, max_row=risk_row + 10)
    cats = Reference(ws, min_col=14, min_row=risk_row + 1, max_row=risk_row + 10)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.legend = None
    chart3.dataLabels = DataLabelList()
    chart3.dataLabels.showVal = True
    ws.add_chart(chart3, "J53")


def apply_number_format(cell, metric_name: Any) -> None:
    name = normalize_text(metric_name)
    if "avance" in name or "probabilidad" in name or "riesgo de atraso" in name:
        cell.number_format = "0%"
    elif any(token in name for token in ("spi", "cpi", "tcpi", "irpc")):
        cell.number_format = "0.00"
    elif any(token in name for token in ("bac", "eac", "etc", "vac", "sv", "cv", "forecast financiero")):
        cell.number_format = '$#,##0;[Red]($#,##0)'
    else:
        cell.number_format = "General"


def apply_conditional_formatting(ws) -> None:
    ws.conditional_formatting.add("C20:C26", CellIsRule(operator="equal", formula=['"Rojo"'], fill=PatternFill("solid", fgColor=PALETTE["red_light"])))
    ws.conditional_formatting.add("G20:G26", CellIsRule(operator="equal", formula=['"Rojo"'], fill=PatternFill("solid", fgColor=PALETTE["red_light"])))
    ws.conditional_formatting.add("C20:C26", CellIsRule(operator="equal", formula=['"Amarillo"'], fill=PatternFill("solid", fgColor=PALETTE["amber_light"])))
    ws.conditional_formatting.add("G20:G26", CellIsRule(operator="equal", formula=['"Amarillo"'], fill=PatternFill("solid", fgColor=PALETTE["amber_light"])))


def apply_borders(ws) -> None:
    thin = Side(style="thin", color="E5E7EB")
    for row in ws.iter_rows(min_row=1, max_row=156, min_col=1, max_col=12):
        for cell in row:
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if cell.alignment is None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in range(1, 157):
        for col in range(1, 13):
            cell = ws.cell(row, col)
            if cell.value is not None and not cell.font:
                cell.font = Font(color=PALETTE["black"], size=10)


def default_workbook_path(folder: Path) -> Path:
    candidates = [
        path
        for path in folder.glob("*.xlsx")
        if not path.name.startswith("~$")
        and OUTPUT_SUFFIX not in path.stem
        and not path.name.startswith(".")
    ]
    if not candidates:
        raise FileNotFoundError("No se encontró ningún archivo .xlsx de entrada en el folder.")
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0]


def output_path_for(input_path: Path, explicit_output: str | None) -> Path:
    if explicit_output:
        return Path(explicit_output).expanduser().resolve()
    return input_path.parent / "outputs" / f"{input_path.stem}{OUTPUT_SUFFIX}{input_path.suffix}"


def generate(input_path: Path, output_path: Path, prompt_folder: Path, use_openai: bool = False) -> Path:
    workbook = load_workbook(input_path)
    table = detect_source_table(workbook)
    rows = read_rows(workbook, table)
    prompt_text = read_prompt_text(prompt_folder)
    write_dashboard(workbook, table, rows, prompt_text, use_openai=use_openai)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera automáticamente un dashboard PMO/EVM en Excel desde una exportación de MS Project."
    )
    parser.add_argument("excel", nargs="?", help="Archivo .xlsx de entrada. Si se omite, toma el .xlsx más reciente del folder.")
    parser.add_argument("-o", "--output", help="Ruta del archivo .xlsx de salida.")
    parser.add_argument(
        "--use-openai",
        "-use-openai",
        action="store_true",
        help="Usa OpenAI API para mejorar la narrativa ejecutiva si OPENAI_API_KEY está configurada.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    folder = Path.cwd()
    env_path = load_env_file(folder)
    input_path = Path(args.excel).expanduser().resolve() if args.excel else default_workbook_path(folder)
    output_path = output_path_for(input_path, args.output)
    actual_output_path = generate(input_path, output_path, folder, use_openai=args.use_openai)
    if env_path:
        print(f"Configuracion cargada desde: {env_path}")
    print(f"Dashboard generado: {actual_output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
